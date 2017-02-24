import requests
import re
import argparse

from random import sample
from io import BytesIO
from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_courses_list_html():
    website_address = "https://www.coursera.org/sitemap~www~courses.xml"
    return requests.get(website_address).content
    

def get_random_courses(html, courses_count):
    tree = etree.parse(BytesIO(html))
    root = tree.getroot()
    courses_urls = sample([url[0].text for url in root], courses_count)
    
    courses_list = []
    for url in courses_urls:
        course = get_course_info(get_course_html(url),url)
        if course is not None:
            courses_list.append(course)
    return courses_list


def get_course_html(course_slug):
    course_html = requests.get(course_slug)
    return course_html.text


def get_course_info(raw_html, corse_url):
    soup = BeautifulSoup(raw_html, "lxml")
    number_weeks = len(soup.findAll(class_="week"))
    course_info = {}
    course_info['title'] = soup.find(class_='title display-3-text').text
    course_info['language'] = soup.find(class_='language-info').text
    course_info['week'] = number_weeks if number_weeks != 0 else "No info"
    course_info['course_url'] = corse_url
    course_info['starts'] = get_starts(soup)
    course_info['rating'] = get_rating(soup)
    return course_info


def get_starts(soup):
    start_date = soup.find('div', {'class':'startdate'})
    if start_date:
        return start_date.text
    return "No info"
    

def get_rating(soup):
    result_tag = soup.find(class_='ratings-text bt3-hidden-xs')
    if result_tag is not None:
        return re.search(r"\d+.\d+", result_tag.text).group(0)
    return "No info"
    

def output_courses_info_to_xlsx(filepath, courses, keys):
    wb = Workbook()
    ws = wb.active
    
    for column, key in enumerate(keys, start=1):
        ws.cell(row=1, column=column, value=key)

    for row, item in enumerate(courses, start=1):
        for column, key in enumerate(keys, start=1):
            ws.cell(row=row, column=column, value=item[key])

    wb.save(filepath)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='The script obtains Coursera courses information and unloads it to the xlsx-file')
    parser.add_argument('-fn', '--xlsx_file_name', default="Result.xlsx", 
                        help='xlsx-file name')
    parser.add_argument('-c', '--count', default=20, 
                        help='Count courses', type=int)
    args = parser.parse_args()
    
    html = get_courses_list_html()
    courses_list = get_random_courses(html, args.count)
    keys = ['title', 'starts', 'language', 'week', 'rating', 'course_url']
    output_courses_info_to_xlsx(args.xlsx_file_name, courses_list, keys)
    print("Finish")
